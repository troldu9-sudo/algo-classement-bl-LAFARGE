"""Generation du tableur de suivi."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import Config
from .modeles import Resultat

EUROS = '#,##0.00\\ "€"'
METRES_CUBES = '#,##0.000\\ "m³"'
QUANTITE = "#,##0.000"
POURCENT = '0.00\\ "%"'
DATE = "DD/MM/YYYY"

_FOND_ENTETE = PatternFill("solid", fgColor="1F3864")
_POLICE_ENTETE = Font(color="FFFFFF", bold=True)
_POLICE_LIEN = Font(color="0563C1", underline="single")
_FOND_ALERTE = PatternFill("solid", fgColor="FCE4E4")


class Colonne:
    """Definition d'une colonne du tableur : titre, largeur, format, source de la valeur."""

    def __init__(self, titre: str, largeur: int, format_nombre: str = "", lien: bool = False):
        self.titre = titre
        self.largeur = largeur
        self.format_nombre = format_nombre
        self.lien = lien


def _ecrire_entetes(feuille: Worksheet, colonnes: list[Colonne]) -> None:
    for index, colonne in enumerate(colonnes, start=1):
        cellule = feuille.cell(row=1, column=index, value=colonne.titre)
        cellule.fill = _FOND_ENTETE
        cellule.font = _POLICE_ENTETE
        cellule.alignment = Alignment(vertical="center", wrap_text=True)
        feuille.column_dimensions[get_column_letter(index)].width = colonne.largeur
    feuille.row_dimensions[1].height = 30


def _ecrire_lignes(feuille: Worksheet, colonnes: list[Colonne], lignes: list[list]) -> None:
    for numero_ligne, valeurs in enumerate(lignes, start=2):
        for index, (colonne, valeur) in enumerate(zip(colonnes, valeurs), start=1):
            if colonne.lien and isinstance(valeur, Path):
                cellule = feuille.cell(row=numero_ligne, column=index, value=valeur.name)
                cellule.hyperlink = str(valeur)
                cellule.font = _POLICE_LIEN
                continue
            cellule = feuille.cell(row=numero_ligne, column=index, value=valeur)
            if colonne.format_nombre and valeur is not None:
                cellule.number_format = colonne.format_nombre


def _finaliser(feuille: Worksheet, colonnes: list[Colonne], nombre_lignes: int) -> None:
    feuille.freeze_panes = "A2"
    if nombre_lignes:
        derniere = get_column_letter(len(colonnes))
        feuille.auto_filter.ref = f"A1:{derniere}{nombre_lignes + 1}"


def _feuille(classeur: Workbook, titre: str, colonnes: list[Colonne], lignes: list[list]) -> Worksheet:
    feuille = classeur.create_sheet(titre)
    _ecrire_entetes(feuille, colonnes)
    _ecrire_lignes(feuille, colonnes, lignes)
    _finaliser(feuille, colonnes, len(lignes))
    return feuille


def _synthese(classeur: Workbook, resultat: Resultat) -> None:
    colonnes = [
        Colonne("N° facture", 14),
        Colonne("Date facture", 13, DATE),
        Colonne("Echeance", 13, DATE),
        Colonne("Agence Lafarge", 20),
        Colonne("Client", 32),
        Colonne("Code client", 12),
        Colonne("Chantiers", 34),
        Colonne("Nb bons", 9),
        Colonne("Volume", 13, METRES_CUBES),
        Colonne("Total HT", 14, EUROS),
        Colonne("Total TVA", 14, EUROS),
        Colonne("Total TTC", 14, EUROS),
        Colonne("PDF facture", 30, lien=True),
    ]
    lignes = [
        [
            f.numero,
            f.date_facture,
            f.date_echeance,
            f.agence,
            f.nom_client,
            f.code_client,
            f.chantiers_intitules,
            len(f.livraisons),
            f.volume_total,
            f.total_ht,
            f.total_tva,
            f.total_ttc,
            f.pdf_a_referencer,
        ]
        for f in resultat.factures
    ]
    _feuille(classeur, "Synthese", colonnes, lignes)


def _detail_bons(classeur: Workbook, resultat: Resultat) -> None:
    colonnes = [
        Colonne("N° facture", 14),
        Colonne("Date facture", 13, DATE),
        Colonne("N° Bon", 16),
        Colonne("Variante", 10),
        Colonne("Date livraison", 13, DATE),
        Colonne("Site expediteur", 16),
        Colonne("Code chantier", 13),
        Colonne("Chantier", 28),
        Colonne("Adresse chantier", 34),
        Colonne("Designation", 44),
        Colonne("Quantite", 12, QUANTITE),
        Colonne("Unite", 8),
        Colonne("Prix unitaire", 14, EUROS),
        Colonne("Montant HT", 14, EUROS),
        Colonne("Code TVA", 9),
        Colonne("Poids eq CO2", 12),
        Colonne("Reduction CO2", 13),
        Colonne("Classe GWR", 11),
        Colonne("PDF facture", 28, lien=True),
        Colonne("PDF bon", 28, lien=True),
    ]
    adresses = {
        (f.numero, c.code): ", ".join(c.adresse) for f in resultat.factures for c in f.chantiers
    }
    lignes = []
    for facture in resultat.factures:
        for l in facture.livraisons:
            lignes.append(
                [
                    facture.numero,
                    facture.date_facture,
                    l.n_bon_brut,
                    l.variante,
                    l.date_livraison,
                    l.site_expediteur,
                    l.chantier_code,
                    l.chantier_nom,
                    adresses.get((facture.numero, l.chantier_code), ""),
                    l.libelle,
                    l.quantite,
                    l.unite,
                    l.prix_unitaire,
                    l.montant_ht,
                    l.code_tva,
                    l.poids_co2,
                    l.reduction_co2,
                    l.classe_gwr,
                    facture.pdf_a_referencer,
                    l.pdf_bon,
                ]
            )
    _feuille(classeur, "Detail BL", colonnes, lignes)


def _annexes(classeur: Workbook, resultat: Resultat) -> None:
    colonnes = [
        Colonne("N° facture", 14),
        Colonne("Date facture", 13, DATE),
        Colonne("Code chantier", 13),
        Colonne("Chantier", 28),
        Colonne("Libelle", 52),
        Colonne("Quantite", 12, QUANTITE),
        Colonne("Unite", 8),
        Colonne("Prix unitaire", 14, EUROS),
        Colonne("Montant HT", 14, EUROS),
        Colonne("Code TVA", 9),
    ]
    lignes = [
        [
            f.numero,
            f.date_facture,
            a.chantier_code,
            a.chantier_nom,
            a.libelle,
            a.quantite,
            a.unite,
            a.prix_unitaire,
            a.montant_ht,
            a.code_tva,
        ]
        for f in resultat.factures
        for a in f.annexes
    ]
    _feuille(classeur, "Lignes annexes", colonnes, lignes)


def _a_verifier(classeur: Workbook, resultat: Resultat) -> None:
    colonnes = [
        Colonne("Categorie", 16),
        Colonne("Facture", 14),
        Colonne("Detail", 86),
        Colonne("Fichier", 34),
    ]
    lignes = [[a.categorie, a.facture, a.detail, a.fichier] for a in resultat.anomalies]
    feuille = _feuille(classeur, "A verifier", colonnes, lignes)
    for numero_ligne in range(2, len(lignes) + 2):
        for index in range(1, len(colonnes) + 1):
            feuille.cell(row=numero_ligne, column=index).fill = _FOND_ALERTE
    if not lignes:
        feuille.cell(row=2, column=1, value="Aucune anomalie detectee.")


def _parametres(classeur: Workbook, resultat: Resultat, config: Config) -> None:
    colonnes = [Colonne("Reglage", 30), Colonne("Valeur", 90)]
    volume = sum(f.volume_total or 0 for f in resultat.factures)
    montant = sum(f.total_ht or 0 for f in resultat.factures)
    lignes = [
        ["Date d'execution", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["Dossier analyse", str(config.dossier)],
        ["Factures lues", len(resultat.factures)],
        ["Bons de livraison", sum(len(f.livraisons) for f in resultat.factures)],
        ["Bons rapproches d'un PDF", sum(1 for f in resultat.factures for l in f.livraisons if l.pdf_bon)],
        ["Volume total", f"{volume:.3f} m³"],
        ["Total HT cumule", f"{montant:.2f} €"],
        ["Anomalies", len(resultat.anomalies)],
        ["Decoupage des PDF groupes", "oui" if config.decouper_pdf_groupes else "non"],
        ["Rapprochement des bons", "oui" if config.apparier_bons_de_livraison else "non"],
        ["Types de page ignores", ", ".join(config.types_page_ignores)],
        ["Tolerance sur les totaux", f"{config.tolerance_totaux:.2f} €"],
        ["Tolerance de regroupement des lignes", f"{config.tolerance_ligne:.1f} points"],
    ]
    lignes += [
        [f"Colonne « {c.nom} »", f"repere {c.repere}, de {c.mini:.0f} a {c.maxi:.0f} points"]
        for c in config.colonnes
    ]
    _feuille(classeur, "Parametres", colonnes, lignes)


def nom_fichier(config: Config, jour: date | None = None) -> str:
    jour = jour or date.today()
    return f"{config.prefixe_tableur}_{jour.isoformat()}.xlsx"


def ecrire(resultat: Resultat, config: Config, cible: Path) -> Path:
    """Ecrit le classeur complet et renvoie son chemin."""
    classeur = Workbook()
    classeur.remove(classeur.active)
    _synthese(classeur, resultat)
    _detail_bons(classeur, resultat)
    _annexes(classeur, resultat)
    _a_verifier(classeur, resultat)
    _parametres(classeur, resultat, config)
    cible.parent.mkdir(parents=True, exist_ok=True)
    classeur.save(str(cible))
    return cible
