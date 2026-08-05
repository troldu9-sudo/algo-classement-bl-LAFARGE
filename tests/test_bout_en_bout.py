"""Chaine complete : lecture du dossier, decoupage, rapprochement, tableur."""

from __future__ import annotations

from openpyxl import load_workbook

from src import excel, main


def _executer(config):
    resultat = main.analyser(config)
    cible = config.dossier / excel.nom_fichier(config)
    excel.ecrire(resultat, config, cible)
    return resultat, cible


def test_le_dossier_entier_est_traite(config_demo):
    resultat, _ = _executer(config_demo)

    numeros = sorted(f.numero for f in resultat.factures)
    assert numeros == [
        "1900200711",
        "1900200712",
        "1900200713",
        "1900300001",
        "1900300002",
        "1900300003",
    ]
    assert sum(len(f.livraisons) for f in resultat.factures) == 13


def test_les_pdf_groupes_sont_decoupes(config_demo):
    resultat, _ = _executer(config_demo)

    decoupes = sorted(p.name for p in config_demo.dossier_decoupe.glob("*.pdf"))
    assert decoupes == [
        "FACTURE_1900300001_2026-05-10.pdf",
        "FACTURE_1900300002_2026-05-11.pdf",
        "FACTURE_1900300003_2026-05-12.pdf",
    ]
    groupees = [f for f in resultat.factures if f.numero.startswith("19003")]
    assert all(f.pdf_decoupe is not None for f in groupees)


def test_les_factures_isolees_ne_sont_pas_recopiees(config_demo):
    """Un PDF qui ne contient qu'une facture est laisse tel quel."""
    resultat, _ = _executer(config_demo)

    simple = next(f for f in resultat.factures if f.numero == "1900200711")
    assert simple.pdf_decoupe is None
    assert simple.pdf_a_referencer.name == "Facture_1900200711.pdf"


def test_aucun_fichier_d_origine_n_est_modifie(config_demo):
    avant = {
        p.name: p.read_bytes() for p in config_demo.dossier.glob("*.pdf")
    }

    _executer(config_demo)

    apres = {p.name: p.read_bytes() for p in config_demo.dossier.glob("*.pdf")}
    assert apres == avant


def test_les_bons_sont_rapproches(config_demo):
    resultat, _ = _executer(config_demo)

    rapproches = {
        l.n_bon: l.pdf_bon.name
        for f in resultat.factures
        for l in f.livraisons
        if l.pdf_bon
    }
    assert rapproches == {
        "293982": "BL_CE293982.pdf",
        "294100": "Bon-de-livraison-juillet.pdf",
    }


def test_le_pdf_scanne_est_signale(config_demo):
    resultat, _ = _executer(config_demo)

    scannes = [a for a in resultat.anomalies if a.categorie == "PDF scanne"]
    assert [a.fichier for a in scannes] == ["Bon_scanne.pdf"]


def test_aucune_anomalie_de_total(config_demo):
    """Les seules anomalies attendues concernent les bons sans PDF et le fichier scanne."""
    resultat, _ = _executer(config_demo)

    assert [a.categorie for a in resultat.anomalies if a.categorie == "Total"] == []


def test_deux_executions_donnent_le_meme_resultat(config_demo):
    premier, _ = _executer(config_demo)
    empreintes = {p.name: p.read_bytes() for p in config_demo.dossier_decoupe.glob("*.pdf")}

    second, _ = _executer(config_demo)

    assert len(second.factures) == len(premier.factures)
    assert len(second.anomalies) == len(premier.anomalies)
    apres = {p.name: p.read_bytes() for p in config_demo.dossier_decoupe.glob("*.pdf")}
    assert apres == empreintes


def test_le_tableur_contient_les_cinq_onglets(config_demo):
    _, cible = _executer(config_demo)

    classeur = load_workbook(cible)
    assert classeur.sheetnames == [
        "Synthese",
        "Detail BL",
        "Lignes annexes",
        "A verifier",
        "Parametres",
    ]


def test_l_onglet_detail_liste_tous_les_bons(config_demo):
    _, cible = _executer(config_demo)

    feuille = load_workbook(cible)["Detail BL"]
    assert feuille.max_row == 14  # treize bons plus la ligne d'en-tete
    entetes = [c.value for c in feuille[1]]
    assert entetes[:3] == ["N° facture", "Date facture", "N° Bon"]


def test_le_tableur_pointe_vers_les_pdf(config_demo):
    _, cible = _executer(config_demo)

    feuille = load_workbook(cible)["Detail BL"]
    colonne_bon = [c.value for c in feuille[1]].index("PDF bon") + 1
    liens = [
        feuille.cell(row=ligne, column=colonne_bon).hyperlink
        for ligne in range(2, feuille.max_row + 1)
    ]
    assert sum(1 for lien in liens if lien is not None) == 2


def test_la_synthese_totalise_chaque_facture(config_demo):
    _, cible = _executer(config_demo)

    feuille = load_workbook(cible)["Synthese"]
    entetes = [c.value for c in feuille[1]]
    ligne = next(
        r for r in feuille.iter_rows(min_row=2, values_only=True) if r[0] == "1900200711"
    )
    assert ligne[entetes.index("Total HT")] == 401.12
    assert ligne[entetes.index("Total TTC")] == 481.34
    assert ligne[entetes.index("Nb bons")] == 1


def test_dossier_vide(config_demo, tmp_path):
    config_demo.dossier = tmp_path / "vide"
    config_demo.dossier.mkdir()

    resultat = main.analyser(config_demo)

    assert resultat.factures == []
    assert resultat.anomalies == []
