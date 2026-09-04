"""Le tableur reste ouvert dans Excel : Windows le verrouille.

C'est le cas le plus frequent en usage reel, puisque l'outil ouvre lui-meme le tableur
a la fin : au lancement suivant, il se heurte a son propre fichier. Le travail d'une
execution ne doit jamais etre perdu pour cette raison.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src import excel, main
from src.modeles import Resultat


def _bloquer(noms_bloques: set[str], monkeypatch: pytest.MonkeyPatch) -> None:
    """Simule le verrou Windows : ces noms de fichiers refusent l'ecriture."""
    vrai_save = Workbook.save

    def save(self, filename):
        if Path(filename).name in noms_bloques:
            raise PermissionError(13, "Permission denied", str(filename))
        return vrai_save(self, filename)

    monkeypatch.setattr(Workbook, "save", save)


def test_verrouille_est_faux_sur_un_fichier_absent(tmp_path):
    assert not excel.verrouille(tmp_path / "rien.xlsx")


def test_verrouille_est_faux_sur_un_fichier_ecrivable(tmp_path):
    fichier = tmp_path / "libre.xlsx"
    fichier.write_bytes(b"contenu")
    assert not excel.verrouille(fichier)


def test_le_tableur_est_ecrit_a_cote_quand_la_cible_est_verrouillee(
    config_demo, monkeypatch
):
    resultat = main.analyser(config_demo)
    cible = config_demo.dossier / excel.nom_fichier(config_demo)
    _bloquer({cible.name}, monkeypatch)

    ecrit = excel.ecrire(resultat, config_demo, cible)

    assert ecrit.name == f"{cible.stem}_2.xlsx"
    assert ecrit.exists()
    assert load_workbook(ecrit)["Detail BL"].max_row == 14


def test_le_repli_continue_tant_que_les_noms_sont_pris(config_demo, monkeypatch):
    resultat = main.analyser(config_demo)
    cible = config_demo.dossier / excel.nom_fichier(config_demo)
    _bloquer({cible.name, f"{cible.stem}_2.xlsx", f"{cible.stem}_3.xlsx"}, monkeypatch)

    ecrit = excel.ecrire(resultat, config_demo, cible)

    assert ecrit.name == f"{cible.stem}_4.xlsx"


def test_message_clair_si_aucun_nom_n_est_ecrivable(config_demo, monkeypatch):
    """Dossier en lecture seule : l'outil doit expliquer, pas empiler une trace d'erreur."""
    cible = config_demo.dossier / excel.nom_fichier(config_demo)
    monkeypatch.setattr(
        Workbook,
        "save",
        lambda self, filename: (_ for _ in ()).throw(PermissionError(13, "refuse")),
    )

    with pytest.raises(excel.TableurVerrouille) as erreur:
        excel.ecrire(Resultat(), config_demo, cible)

    message = str(erreur.value)
    assert "ouverts dans Excel" in message
    assert "lecture seule" in message


def test_l_outil_rend_un_code_d_erreur_plutot_qu_une_trace(config_demo, monkeypatch):
    monkeypatch.setattr(main, "charger", lambda: config_demo)
    monkeypatch.setattr(
        Workbook,
        "save",
        lambda self, filename: (_ for _ in ()).throw(PermissionError(13, "refuse")),
    )

    assert main.executer([]) == 4


def test_le_tableur_de_repli_est_celui_qui_est_ouvert(config_demo, monkeypatch):
    """L'utilisateur doit voir s'ouvrir le fichier reellement ecrit, pas la cible ratee."""
    cible = config_demo.dossier / excel.nom_fichier(config_demo)
    _bloquer({cible.name}, monkeypatch)
    monkeypatch.setattr(main, "charger", lambda: config_demo)
    ouverts = []
    monkeypatch.setattr(main, "_ouvrir", ouverts.append)

    assert main.executer([]) == 0
    assert [p.name for p in ouverts] == [f"{cible.stem}_2.xlsx"]
